import unittest
import sys
sys.path.append(".")

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment, PatternFill
from openpyxl.styles.fills import FILL_SOLID

from tablepyxl.tablepyxl import string_to_int, get_Tables, document_to_workbook, insert_table_at_cell, table_to_sheet
from tablepyxl.style import style_string_to_dict, style_dict_to_named_style, StyleDict, known_styles


table_one = "<table name='simple table'> " \
            "<tbody> " \
            "<tr> " \
            "<td>A cell</td> " \
            "<td>=1+2</td> " \
            "<td class='TYPE_FORMULA TYPE_INTEGER'>=1+2</td> " \
            "</tr> " \
            "</tbody> " \
            "</table>"

table_two = "<table name='second table'><thead></thead> " \
            "<tbody> " \
            "<tr> " \
            "<td>Another cell</td> " \
            "<td>B1 cell</td> " \
            "</tr> " \
            "<tr></tr>" \
            "</tbody> " \
            "</table>"

table_three = "<table name='Another simple table'><thead></thead> " \
              "<tbody> " \
              "<tr> " \
              "<td>T3 cell</td> " \
              "</tr> " \
              "</tbody> " \
              "</table>"

table_span = "<table name='span table'><thead></thead> " \
             "<tbody> " \
             "<tr> " \
             "<td colspan='3'>A1 through C1 cell</td> " \
             "</tr> " \
             "<tr> " \
             "<td rowspan='4'>A2 through A5 cell</td> " \
             "</tr> " \
             "</tbody> " \
             "</table>"

table_widths = "<table name='width table'><thead></thead> " \
               "<tbody> " \
               "<tr> " \
               "<td>12</td> " \
               "</tr> " \
               "<tr> " \
               "<td colspan='2'>123456789</td> " \
               "</tr> " \
               "<tr> " \
               "<td>1234</td> " \
               "</tr> " \
               "<tr> " \
               "<td>123</td> " \
               "</tr> " \
               "</tbody> " \
               "</table>"

table_whitespace = "<table name='whitespace table'>" \
                   "<thead></thead>" \
                   "<tbody>" \
                   "<tr>" \
                   "<td>   a  bc  <inn1>  d  ef  </inn1>  g  hi   <inn2>   j  k  <inn3>  l  m  </inn3>  n  o  </inn2>  p  </td>" \
                   "</tr>" \
                   "</tbody>" \
                   "</table>"

table_comment = "<table name='comment table'>" \
                "<tr>" \
                "<td><!-- this is a html comment --></td>" \
                "<td>this is not a html comment</td>" \
                "</tr>" \
                "</table>"


class TestTablepyxl(unittest.TestCase):
    """
    Unit tests for tablepyxl.py
    """

    def test_string_to_int(self):
        self.assertEqual(string_to_int('3'), 3)
        self.assertEqual(string_to_int('3.1'), 0)

    def test_get_tables(self):
        doc = table_one
        self.assertEqual(len(get_Tables(doc)), 1)

        doc += doc
        self.assertEqual(len(get_Tables(doc)), 2)

    def test_table_to_sheet(self):
        wb = Workbook()
        table = get_Tables(table_one)
        table_to_sheet(table[0], wb)

        sheet = wb['simple table']  # Get sheet with the title `simple table`
        self.assertEqual(sheet['A1'].value, 'A cell')
        self.assertEqual(sheet['B1'].value, '=1+2')
        self.assertEqual(sheet['B1'].data_type, 's')
        self.assertEqual(sheet['B1'].number_format, 'General')
        self.assertEqual(sheet['C1'].value, '=1+2')
        self.assertEqual(sheet['C1'].data_type, 'f')
        self.assertEqual(sheet['C1'].number_format, '#,##0')

    def test_document_to_workbook(self):
        doc = table_one + table_two
        wb = document_to_workbook(doc)
        self.assertEqual(wb.sheetnames, ['simple table', 'second table'])

        sheet = wb['second table']  # Get sheet with the title `span table`
        self.assertEqual(sheet['B1'].value, 'B1 cell')

        # Add another sheet to the existing workbook
        wb = document_to_workbook(table_three, wb=wb)
        self.assertEqual(wb.sheetnames, ['simple table', 'second table', 'Another simple table'])

    def test_comments(self):
        wb = document_to_workbook(table_comment)
        sheet = wb['comment table']
        self.assertNotIn('this is a html comment', sheet['A1'].value)
        self.assertEqual(sheet['B1'].value, 'this is not a html comment')

    def test_spans(self):
        doc = table_span
        wb = document_to_workbook(doc)
        sheet = wb['span table']  # Get sheet with the title `span table`
        self.assertIn("A1:C1", sheet.merged_cell_ranges)
        self.assertIn("A2:A5", sheet.merged_cell_ranges)

    def test_width(self):
        doc = table_widths
        wb = document_to_workbook(doc)
        sheet = wb['width table']  # Get sheet with the title `width table`
        self.assertEqual(sheet.column_dimensions['A'].width, 6)

    def test_insert_table_at_cell(self):
        wb = Workbook()
        ws = wb.active
        cell = ws['B2']

        table = get_Tables(table_one)
        insert_table_at_cell(table[0], cell)

        self.assertEqual(ws['B2'].value, 'A cell')

    def test_element_whitespace(self):
        doc = table_whitespace
        wb = document_to_workbook(doc)
        sheet = wb['whitespace table']

        self.assertEqual(sheet['A1'].value, 'a  bc\nd  ef\ng  hi\nj  k\nl  m\nn  o\np')


class TestStyle(unittest.TestCase):
    """
    Unit tests for style.py
    """

    def test_style_string_to_dict(self):
        string = 'key: value'
        d = {'key': 'value'}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'key: value;'
        d = {'key': 'value'}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'key: value; k: v'
        d = {'key': 'value', 'k': 'v'}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'invalid; style'
        d = {}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'partially: valid; style'
        d = {'partially': 'valid'}
        self.assertEqual(d, style_string_to_dict(string))

    def test_style_dict_to_style(self):
        d = StyleDict({'font-weight': 'bold'})
        s = NamedStyle(name='Style {}'.format(len(known_styles) + 1), font=Font(bold=True),
                       alignment=Alignment(horizontal='general', vertical=None, wrap_text=False))
        self.assertEqual(style_dict_to_named_style(d), s)

        default_alignment = Alignment(horizontal='general', wrap_text=False)

        d = StyleDict({'color': 'ff0000'})
        s = NamedStyle(name='Style {}'.format(len(known_styles) + 1), font=Font(bold=False, color='ff0000'),
                       alignment=default_alignment)

        self.assertEqual(style_dict_to_named_style(d), s)

        d = StyleDict({'text-align': 'left'})
        s = NamedStyle(name='Style {}'.format(len(known_styles) + 1), alignment=Alignment(horizontal='left', wrap_text=False))

        self.assertEqual(style_dict_to_named_style(d), s)

        d = StyleDict({'background-color': '#ff0000'})
        s = NamedStyle(name='Style {}'.format(len(known_styles) + 1), fill=PatternFill(fill_type=FILL_SOLID, start_color="ff0000"),
                       alignment=default_alignment)
        self.assertEqual(style_dict_to_named_style(d), s)

        # Make sure we reuse a style when it already exists
        known_styles_length = len(known_styles)
        d = StyleDict({'background-color': '#ff0000'})
        style_dict_to_named_style(d)
        self.assertEqual(len(known_styles), known_styles_length)

        # Create new one when it doesn't
        known_styles_length = len(known_styles)
        d = StyleDict({'background-color': '#ff0000', 'class': 'TYPE_NUMERIC'})
        style_dict_to_named_style(d)
        self.assertEqual(len(known_styles), known_styles_length + 1)


    def test_parent(self):
        parent = StyleDict({'parent': 'mother'})
        child = StyleDict({'child': 'daughter'}, parent=parent)

        self.assertEqual(child.get('parent'), 'mother')
        self.assertEqual(sorted(child._keys()), ['child', 'parent'])

        # We should get the definition of child from the child
        # not the parent and child should show up only once in the keys
        parent = StyleDict({'parent': 'mother', 'child': 'son'})
        child = StyleDict({'child': 'daughter'}, parent=parent)
        self.assertEqual(child.get('parent'), 'mother')
        self.assertEqual(child.get('child'), 'daughter')
        self.assertEqual(sorted(child._keys()), ['child', 'parent'])

if __name__ == "__main__":
    unittest.main()
