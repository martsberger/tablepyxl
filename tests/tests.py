import unittest

import sys
sys.path.append("/home/travis/build/martsberger/tablepyxl")
print "path", sys.path
from openpyxl import Workbook
from openpyxl.styles import Font, Style, Alignment, PatternFill
from openpyxl.styles.fills import FILL_SOLID

from tablepyxl.tablepyxl import string_to_int, get_Tables, document_to_workbook, insert_table_at_cell, table_to_sheet
from tablepyxl.style import style_string_to_dict, style_dict_to_Style, StyleDict


table_one = "<table name='simple table'> " \
            "<tbody> " \
            "<tr> " \
            "<td>A cell</td> " \
            "</tr> " \
            "</tbody> " \
            "</table>"

table_two = "<table name='second table'><thead></thead> " \
            "<tbody> " \
            "<tr> " \
            "<td>Another cell</td> " \
            "<td>B1 cell</td> " \
            "</tr> " \
            "<tr></tr>" \
            "</tbody> " \
            "</table>"

table_three = "<table name='Another simple table'><thead></thead> " \
              "<tbody> " \
              "<tr> " \
              "<td>T3 cell</td> " \
              "</tr> " \
              "</tbody> " \
              "</table>"

table_span = "<table name='span table'><thead></thead> " \
             "<tbody> " \
             "<tr> " \
             "<td colspan=3>A1 through C1 cell</td> " \
             "</tr> " \
             "<tr> " \
             "<td rowspan=4>A2 through A5 cell</td> " \
             "</tr> " \
             "</tbody> " \
             "</table>"


class TestTablepyxl(unittest.TestCase):
    """
    Unit tests for tablepyxl.py
    """

    def test_string_to_int(self):
        self.assertEqual(string_to_int('3'), 3)
        self.assertEqual(string_to_int('3.1'), 0)

    def test_get_tables(self):
        doc = table_one
        self.assertEqual(len(get_Tables(doc)), 1)

        doc += doc
        self.assertEqual(len(get_Tables(doc)), 2)

    def test_table_to_sheet(self):
        wb = Workbook()
        table = get_Tables(table_one)
        table_to_sheet(table[0], wb)

        sheet = wb.get_sheet_by_name('simple table')
        self.assertEqual(sheet.cell('A1').value, 'A cell')

    def test_document_to_workbook(self):
        doc = table_one + table_two
        wb = document_to_workbook(doc)
        self.assertEqual(wb.sheetnames, ['simple table', 'second table'])

        sheet = wb.get_sheet_by_name('second table')
        self.assertEqual(sheet.cell('B1').value, 'B1 cell')

        # Add another sheet to the existing workbook
        wb = document_to_workbook(table_three, wb=wb)
        self.assertEqual(wb.sheetnames, ['simple table', 'second table', 'Another simple table'])

    def test_spans(self):
        doc = table_span
        wb = document_to_workbook(doc)
        sheet = wb.get_sheet_by_name('span table')
        self.assertIn("A1:C1", sheet.merged_cell_ranges)
        self.assertIn("A2:A5", sheet.merged_cell_ranges)

    def test_insert_table_at_cell(self):
        wb = Workbook()
        ws = wb.active
        cell = ws.cell('B2')

        table = get_Tables(table_one)
        insert_table_at_cell(table[0], cell)

        self.assertEqual(ws.cell('B2').value, 'A cell')


class TestStyle(unittest.TestCase):
    """
    Unit tests for style.py
    """

    def test_style_string_to_dict(self):
        string = 'key: value'
        d = {'key': 'value'}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'key: value;'
        d = {'key': 'value'}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'key: value; k: v'
        d = {'key': 'value', 'k': 'v'}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'invalid; style'
        d = {}
        self.assertEqual(d, style_string_to_dict(string))

        string = 'partially: valid; style'
        d = {'partially': 'valid'}
        self.assertEqual(d, style_string_to_dict(string))

    def test_style_dict_to_Style(self):
        d = StyleDict({'font-weight': 'bold'})
        s = Style(font=Font(bold=True),
                  alignment=Alignment(horizontal='general', vertical=None, wrap_text=False))
        self.assertEqual(style_dict_to_Style(d), s)

        default_alignment = Alignment(horizontal='general', wrap_text=False)

        d = StyleDict({'color': 'ff0000'})
        s = Style(font=Font(bold=False, color='ff0000'),
                  alignment=default_alignment)
        self.assertEqual(style_dict_to_Style(d), s)

        d = StyleDict({'text-align': 'left'})
        s = Style(alignment=Alignment(horizontal='left', wrap_text=False))
        self.assertEqual(style_dict_to_Style(d), s)

        d = StyleDict({'background-color': '#ff0000'})
        s = Style(fill=PatternFill(fill_type=FILL_SOLID, start_color="ff0000"),
                  alignment=default_alignment)
        self.assertEqual(style_dict_to_Style(d), s)

    def test_parent(self):
        parent = StyleDict({'parent': 'mother'})
        child = StyleDict({'child': 'daughter'}, parent=parent)

        self.assertEqual(child.get('parent'), 'mother')
        self.assertEqual(sorted(child._keys()), ['child', 'parent'])

if __name__ == "__main__":
    unittest.main()
