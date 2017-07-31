# This is where we handle translating css styles into openpyxl styles
# and cascading those from parent to child in the dom.

from openpyxl.cell import Cell
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side, Color
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from openpyxl.styles.colors import BLACK

FORMAT_DATE_MMDDYYYY = 'mm/dd/yyyy'


def colormap(color):
    """
    Convenience for looking up known colors
    """
    cmap = {'black': BLACK}
    return cmap.get(color, color)


def style_string_to_dict(style):
    """
    Convert css style string to a python dictionary
    """
    def clean_split(string, delim):
        return (s.strip() for s in string.split(delim))
    styles = [clean_split(s, ":") for s in style.split(";") if ":" in s]
    return dict(styles)


def get_side(style, name):
    return {'border_style': style.get('border-{}-style'.format(name)),
            'color': colormap(style.get('border-{}-color'.format(name)))}

known_styles = {}


def style_dict_to_named_style(style_dict):
    """
    Change css style (stored in a python dictionary) to openpyxl NamedStyle
    """
    if style_dict not in known_styles:
        # Font
        font = Font(bold=style_dict.get('font-weight') == 'bold',
                    color=style_dict.get_color('color', None),
                    size=style_dict.get('font-size'))

        # Alignment
        alignment = Alignment(horizontal=style_dict.get('text-align', 'general'),
                              vertical=style_dict.get('vertical-align'),
                              wrap_text=style_dict.get('white-space', 'nowrap') == 'normal')

        # Fill
        bg_color = style_dict.get_color('background-color')
        fg_color = style_dict.get_color('foreground-color', Color())
        fill_type = style_dict.get('fill-type')
        if bg_color:
            fill = PatternFill(fill_type=fill_type or FILL_SOLID,
                               start_color=bg_color,
                               end_color=fg_color)
        else:
            fill = PatternFill()

        # Border
        border = Border(left=Side(**get_side(style_dict, 'left')),
                        right=Side(**get_side(style_dict, 'right')),
                        top=Side(**get_side(style_dict, 'top')),
                        bottom=Side(**get_side(style_dict, 'bottom')),
                        diagonal=Side(**get_side(style_dict, 'diagonal')),
                        diagonal_direction=None,
                        outline=Side(**get_side(style_dict, 'outline')),
                        vertical=None,
                        horizontal=None)

        pyxl_style = NamedStyle(name=str(style_dict), font=font, fill=fill, alignment=alignment, border=border)

        known_styles[style_dict] = pyxl_style

    return known_styles[style_dict]


class StyleDict(dict):
    """
    It's like a dictionary, but it looks for items in the parent dictionary
    """
    def __init__(self, *args, **kwargs):
        self.parent = kwargs.pop('parent', None)
        super(StyleDict, self).__init__(*args, **kwargs)

    def __getitem__(self, item):
        if item in self:
            return super(StyleDict, self).__getitem__(item)
        elif self.parent:
            return self.parent[item]
        else:
            raise KeyError('{} not found'.format(item))

    def __hash__(self):
        return hash(tuple([(k, self.get(k)) for k in self._keys()]))

    # Yielding the keys avoids creating unnecessary data structures
    # and happily works with both python2 and python3 where the
    # .keys() method is a dictionary_view in python3 and a list in python2.
    def _keys(self):
        yielded = set()
        for k in self.keys():
            yielded.add(k)
            yield k
        if self.parent:
            for k in self.parent._keys():
                if k not in yielded:
                    yielded.add(k)
                    yield k

    def get(self, k, d=None):
        try:
            return self[k]
        except KeyError:
            return d

    def get_color(self, k, d=None):
        """
        Strip leading # off colors if necessary
        """
        color = self.get(k, d)
        if hasattr(color, 'startswith') and color.startswith('#'):
            return color[1:]
        return color


class Element(object):
    """
    Our base class for representing an html element along with a cascading style.
    The element is created along with a parent so that the StyleDict that we store
    can point to the parent's StyleDict.
    """
    def __init__(self, element, parent=None):
        self.element = element
        parent_style = parent.style_dict if parent else None
        self.style_dict = StyleDict(style_string_to_dict(element.get('style', '')), parent=parent_style)
        self._style_cache = None

    def style(self):
        """
        Turn the css styles for this element into an openpyxl NamedStyle.
        """
        if not self._style_cache:
            self._style_cache = style_dict_to_named_style(self.style_dict)
        return self._style_cache

    def get_dimension(self, dimension_key):
        """
        Extracts the dimension from the style dict of the Element and returns it as a float.
        """
        dimension = self.style_dict.get(dimension_key)
        if dimension:
            if dimension[-2:] in ['px', 'em', 'pt', 'in', 'cm']:
                dimension = dimension[:-2]
            dimension = float(dimension)
        return dimension


class Table(Element):
    """
    The concrete implementations of Elements are semantically named for the types of elements we are interested in.
    This defines a very concrete tree structure for html tables that we expect to deal with. I prefer this compared to
    allowing Element to have an arbitrary number of children and dealing with an abstract element tree.

    """
    def __init__(self, table):
        """
        takes an html table object (from BeautifulSoup)
        """
        super(Table, self).__init__(table)
        table_head = table.find('thead')
        self.head = TableHead(table_head, parent=self) if len(table_head) else None
        table_body = table.find('tbody')
        self.body = TableBody(table_body if len(table_body) else table, parent=self)


class TableHead(Element):
    """
    This class maps to the `<th>` element of the html table.
    """
    def __init__(self, head, parent=None):
        super(TableHead, self).__init__(head, parent=parent)
        self.rows = [TableRow(tr, parent=self) for tr in head.find_all('tr')]


class TableBody(Element):
    """
    This class maps to the `<tbody>` element of the html table.
    """
    def __init__(self, body, parent=None):
        super(TableBody, self).__init__(body, parent=parent)
        self.rows = [TableRow(tr, parent=self) for tr in body.find_all('tr')]


class TableRow(Element):
    """
    This class maps to the `<tr>` element of the html table.
    """
    def __init__(self, tr, parent=None):
        super(TableRow, self).__init__(tr, parent=parent)
        self.cells = [TableCell(cell, parent=self) for cell in tr.find_all('th') + tr.find_all('td')]


class TableCell(Element):
    """
    This class maps to the `<td>` element of the html table.
    """
    CELL_TYPES = {'TYPE_STRING', 'TYPE_FORMULA', 'TYPE_NUMERIC', 'TYPE_BOOL', 'TYPE_CURRENCY',
                  'TYPE_NULL', 'TYPE_INLINE', 'TYPE_ERROR', 'TYPE_FORMULA_CACHE_STRING', 'TYPE_INTEGER'}

    def __init__(self, *args, **kwargs):
        super(TableCell, self).__init__(*args, **kwargs)
        self.value = self.element.get_text(separator="\n", strip=True)

    def data_type(self):
        cell_types = self.CELL_TYPES & set(self.element.get('class', []))
        if cell_types:
            if 'TYPE_FORMULA' in cell_types:
                # Make sure TYPE_FORMULA takes precedence over the other classes in the set.
                cell_type = 'TYPE_FORMULA'
            elif 'TYPE_CURRENCY' in cell_types or 'TYPE_INTEGER' in cell_types:
                cell_type = 'TYPE_NUMERIC'
            else:
                cell_type = cell_types.pop()
        else:
            cell_type = 'TYPE_STRING'
        return getattr(Cell, cell_type)

    def number_format(self):
        if 'TYPE_CURRENCY' in self.element.get('class', []):
            return FORMAT_CURRENCY_USD_SIMPLE
        if 'TYPE_INTEGER' in self.element.get('class', []):
            return '#,##0'
        if 'TYPE_DATE' in self.element.get('class', []):
            return FORMAT_DATE_MMDDYYYY
        if self.data_type() == Cell.TYPE_NUMERIC:
            try:
                int(self.value)
            except ValueError:
                return '#,##0.##'
            else:
                return '#,##0'

    def format(self, cell):
        style = self.style()
        cell.font = style.font
        cell.alignment = style.alignment
        cell.fill = style.fill
        cell.border = style.border
        data_type = self.data_type()
        if data_type:
            cell.data_type = data_type
        number_format = self.number_format()
        if number_format:
            cell.number_format = number_format
