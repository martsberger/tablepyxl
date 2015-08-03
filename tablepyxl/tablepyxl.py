from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
from premailer import Premailer
from style import Table


def get_Tables(doc):
    soup = BS(doc)
    return [Table(table) for table in soup.find_all('table')]


def write_rows(worksheet, elem, row, column=1):
    """
    Writes every tr child element of elem to a row in the worksheet

    returns the next row after all rows are written
    """
    initial_column = column
    for table_row in elem.rows:
        for table_cell in table_row.cells:
            cell = worksheet.cell(row=row, column=column, value=table_cell.value)
            table_cell.format(cell)
            if worksheet.column_dimensions.values()[column - 1].width < len(cell.value):
                worksheet.column_dimensions.values()[column - 1].width = len(cell.value)
            column += 1
        row += 1
        column = initial_column
    return row


def table_to_sheet(table, wb):
    """
    Takes a table and workbook and writes the table to a new sheet.
    The sheet title will be the same as the table attribute name.
    """
    ws = wb.create_sheet(title=table.element.get('name', None))
    insert_table(table, ws, 1, 1)


def document_to_workbook(doc, wb=None, base_url=None):
    """
    Takes a string representation of an html document and writes one sheet for
    every table in the document.

    The workbook is returned
    """
    if not wb:
        wb = Workbook()
    wb.remove_sheet(wb.active)
    inline_styles_doc = Premailer(doc, base_url=base_url, remove_classes=False).transform()
    tables = get_Tables(inline_styles_doc)

    for table in tables:
        table_to_sheet(table, wb)

    return wb


def document_to_xl(doc, filename, base_url=None):
    """
    Takes a string representation of an html document and writes one sheet for
    every table in the document. The workbook is written out to a file called filename
    """
    wb = document_to_workbook(doc, base_url=base_url)
    wb.save(filename)


def insert_table(table, worksheet, column, row):
    if table.head:
        row = write_rows(worksheet, table.head, row, column)
    if table.body:
        row = write_rows(worksheet, table.body, row, column)


def insert_table_at_cell(table, cell):
    """
    Inserts a table at the location of an openpyxl Cell object.
    """
    ws = cell.parent
    column, row = cell.column, cell.row
    insert_table(table, ws, column, row)
