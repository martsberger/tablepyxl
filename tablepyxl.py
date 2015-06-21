from bs4 import BeautifulSoup as BS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Style
from premailer import transform


def get_tables(doc):
    """
    Turn a string containing html into a list of tables
    """
    soup = BS(doc)
    return soup.find_all('table')


def get_Tables(doc):
    soup = BS(doc)
    return [Table(table) for table in soup.find_all('table')]


def write_rows(worksheet, elem, row, column=1):
    """
    Writes every tr child element of elem to a row in the worksheet
    elem could be a thead or tbody, so we write each th and td to a cell.

    returns the next row after all rows are written
    """
    initial_column = column
    for table_row in elem.rows:
        for table_cell in table_row.cells:
            cell = worksheet.cell(row=row, column=column)
            cell.value = table_cell.element.text
            style = table_cell.style()
            cell.font = style.font
            cell.alignment = style.alignment
            column += 1
        row += 1
        column = initial_column
    return row


def table_to_sheet(table, wb):
    """
    Takes a table and workbook and writes the table to a new sheet.
    The sheet title will be the same as the table attribute name.
    """
    ws = wb.create_sheet(title=table.element.get('name', None))
    insert_table(table, ws, 1, 1)


def document_to_xl(doc, filename):
    """
    Takes a string representation of an html document and writes one sheet for
    every table in the document. The workbook is written out to a file called filename
    """
    wb = Workbook()
    wb.remove_sheet(wb.active)
    inline_styles_doc = transform(doc)
    tables = get_Tables(inline_styles_doc)

    for table in tables:
        table_to_sheet(table, wb)

    wb.save(filename)


def insert_table(table, worksheet, column, row):
    row = write_rows(worksheet, table.head, row, column)
    row = write_rows(worksheet, table.body, row, column)


def insert_table_at_cell(table, cell):
    """
    Inserts a table at the location of an openpyxl Cell object.
    """
    ws = cell.parent
    column, row = cell.column, cell.row
    insert_table(table, ws, column, row)


def style_string_to_dict(style):
    def clean_split(string, delim):
        return (s.strip() for s in string.split(delim))
    styles = [clean_split(s, ":") for s in style.split(";") if ":" in s]
    return dict(styles)


def style_dict_to_Style(style):
    """
    change css style to openpyxl Style
    """
    # Font
    font_kwargs = {'bold': style.get('font-weight') == 'bold'}
    font = Font(**font_kwargs)

    # Alignment
    alignment_kwargs = {'horizontal': style.get('text-align', 'general')}
    alignment = Alignment(**alignment_kwargs)

    pyxl_style = Style(font=font, alignment=alignment)

    return pyxl_style


class Element(object):
    def __init__(self, element, parent=None):
        self.element = element
        self.parent = parent
        self.style_dict = style_string_to_dict(element.get('style', ''))
        self._style_cache = None

    def get_style(self, key):
        if key in self.style_dict:
            return self.style_dict[key]
        if self.parent:
            return self.parent.get_style(key)
        return None

    def style(self):
        """
        Turn the css styles for this element into an openpyxl Style
        """
        if not self._style_cache:
            self._style_cache = style_dict_to_Style(self.style_dict)
        return self._style_cache


class Table(Element):
    def __init__(self, table):
        """
        takes an html table object (from BeautifulSoup)
        """
        super(Table, self).__init__(table)
        self.head = TableHead(table.thead, parent=self)
        self.body = TableBody(table.tbody, parent=self)


class TableHead(Element):
    def __init__(self, head, parent=None):
        super(TableHead, self).__init__(head, parent=parent)
        self.rows = [TableRow(tr, parent=self) for tr in head.find_all('tr')]


class TableBody(Element):
    def __init__(self, body, parent=None):
        super(TableBody, self).__init__(body, parent=parent)
        self.rows = [TableRow(tr, parent=self) for tr in body.find_all('tr')]


class TableRow(Element):
    def __init__(self, tr, parent=None):
        super(TableRow, self).__init__(tr, parent=parent)
        self.cells = [TableCell(cell, parent=self) for cell in tr.find_all('th') + tr.find_all('td')]


class TableCell(Element):
    pass
